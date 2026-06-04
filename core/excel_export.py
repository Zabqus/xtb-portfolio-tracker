"""
Eksport portfela do sformatowanego Excel (.xlsx) — arkusze: Portfolio, Historia, Analiza.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from core.allocation import REGION_ORDER, aggregate_breakdown, enrich_portfolio_allocation
from core.analyzer import portfolio_summary
from core.closed_analysis import closed_positions_summary
from core.cost_basis import get_current_cost_basis
from core.importer import XTBReport
from core.timeline import build_portfolio_timeline
from core.trade_analytics import TradeAnalyticsSummary, compute_trade_analytics
from core.transactions import parse_cash_operations_trades

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


class ExcelExportError(RuntimeError):
    """Błąd generowania pliku Excel."""


def default_excel_filename() -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return f"xtb_portfolio_{stamp}.xlsx"


def _autosize_columns(ws, min_width: int = 10, max_width: int = 42) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)) + 2, max_width))
        ws.column_dimensions[letter].width = max_len


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _write_section_title(ws, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    return row + 1


def _write_key_values(ws, start_row: int, items: list[tuple[str, object]]) -> int:
    row = start_row
    for label, value in items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row + 1


def _prepare_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.tz_localize(None)
    return out


def _write_dataframe(
    ws,
    df: pd.DataFrame,
    start_row: int,
    *,
    table_title: str | None = None,
) -> int:
    row = start_row
    if table_title:
        row = _write_section_title(ws, row, table_title)

    if df is None or df.empty:
        ws.cell(row=row, column=1, value="Brak danych")
        return row + 2

    prepared = _prepare_df_for_excel(df)
    headers = list(prepared.columns)
    for c_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=c_idx, value=header)
    _style_header_row(ws, row, len(headers))
    row += 1

    for record in dataframe_to_rows(prepared, index=False, header=False):
        for c_idx, value in enumerate(record, start=1):
            cell = ws.cell(row=row, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if isinstance(value, float):
                cell.number_format = "#,##0.00"
        row += 1
    return row + 1


def _portfolio_positions_df(analyzed: pd.DataFrame, currency: str) -> pd.DataFrame:
    cols = [
        "ticker_xtb",
        "ticker_yahoo",
        "currency",
        "quantity",
        "avg_price",
        "market_price",
        "position_cost",
        "market_value",
        "pnl",
        "roi_pct",
    ]
    available = [c for c in cols if c in analyzed.columns]
    df = analyzed[available].sort_values("market_value", ascending=False, na_position="last")
    rename = {
        "ticker_xtb": "Ticker XTB",
        "ticker_yahoo": "Ticker Yahoo",
        "currency": "Waluta",
        "quantity": "Ilość",
        "avg_price": "Śr. cena",
        "market_price": "Cena rynkowa",
        "position_cost": f"Koszt ({currency})",
        "market_value": f"Wartość ({currency})",
        "pnl": f"PnL ({currency})",
        "roi_pct": "ROI %",
    }
    return df.rename(columns={k: v for k, v in rename.items() if k in df.columns})


def _build_portfolio_sheet(ws, report: XTBReport, analyzed: pd.DataFrame, summary: dict) -> None:
    currency = str(summary["display_currency"])
    ws.cell(row=1, column=1, value="XTB Portfolio Tracker — Portfolio").font = TITLE_FONT
    row = 3
    meta = [
        ("Plik źródłowy", report.filename),
        ("Wygenerowano (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")),
        ("Waluta konta", summary["account_currency"]),
        ("Waluta wyświetlania", currency),
    ]
    if report.is_merged and report.account_labels:
        meta.append(("Konta", ", ".join(report.account_labels)))
        if report.source_filenames:
            for label, fn in report.source_filenames.items():
                meta.append((f"Plik {label}", fn))
    elif report.account_number:
        meta.append(("Numer konta", report.account_number))
    row = _write_key_values(ws, row, meta)

    row = _write_section_title(ws, row, "Podsumowanie")
    row = _write_key_values(
        ws,
        row,
        [
            ("Całkowita wartość", round(summary["total_value"], 2)),
            ("Łączny koszt", round(summary["total_cost"], 2)),
            ("Łączny PnL", round(summary["total_pnl"], 2)),
            ("ROI %", round(summary["total_roi_pct"], 2)),
            ("Liczba pozycji", len(analyzed)),
        ],
    )

    positions = _portfolio_positions_df(analyzed, currency)
    _write_dataframe(ws, positions, row, table_title="Otwarte pozycje")
    _autosize_columns(ws)
    ws.freeze_panes = "A4"


def _build_historia_sheet(
    ws,
    report: XTBReport,
    timeline: pd.DataFrame | None,
    trades: pd.DataFrame | None,
    round_trips: pd.DataFrame | None,
    closed: pd.DataFrame | None,
) -> None:
    ws.cell(row=1, column=1, value="Historia").font = TITLE_FONT
    row = 3

    if timeline is not None and not timeline.empty:
        tl = timeline.copy()
        keep = [c for c in ("date", "market_value", "cost_basis", "unrealized_pnl") if c in tl.columns]
        tl = tl[keep].rename(
            columns={
                "date": "Data",
                "market_value": "Wartość rynkowa",
                "cost_basis": "Baza kosztowa",
                "unrealized_pnl": "PnL niezrealizowany",
            }
        )
        row = _write_dataframe(ws, tl, row, table_title="Timeline portfela")
    else:
        row = _write_section_title(ws, row, "Timeline portfela")
        ws.cell(row=row, column=1, value="Brak (wymagany arkusz Cash Operations)")
        row += 2

    if trades is not None and not trades.empty:
        tr = trades.rename(
            columns={
                "trade_time": "Czas (UTC)",
                "ticker_xtb": "Ticker XTB",
                "ticker_yahoo": "Ticker Yahoo",
                "side": "Strona",
                "quantity": "Ilość",
                "price": "Cena",
                "amount": "Kwota",
                "operation_type": "Typ",
                "comment": "Komentarz",
            }
        )
        drop = [c for c in tr.columns if c in ("trade_date",)]
        tr = tr.drop(columns=drop, errors="ignore")
        row = _write_dataframe(ws, tr, row, table_title="Transakcje (Cash Operations)")
    else:
        row = _write_section_title(ws, row, "Transakcje (Cash Operations)")
        ws.cell(row=row, column=1, value="Brak danych")
        row += 2

    if round_trips is not None and not round_trips.empty:
        rt = round_trips.rename(
            columns={
                "ticker_xtb": "Ticker",
                "ticker_yahoo": "Yahoo",
                "open_time": "Otwarcie",
                "close_time": "Zamknięcie",
                "quantity": "Ilość",
                "open_price": "Cena wejścia",
                "close_price": "Cena wyjścia",
                "holding_days": "Dni",
                "realized_pnl": "PnL",
                "pnl_pct": "ROI %",
                "is_win": "Wygrana",
            }
        )
        row = _write_dataframe(ws, rt, row, table_title="Round-tripy FIFO")
    else:
        row = _write_section_title(ws, row, "Round-tripy FIFO")
        ws.cell(row=row, column=1, value="Brak zamkniętych round-tripów")
        row += 2

    if closed is not None and not closed.empty:
        cl = closed.rename(
            columns={
                "ticker_xtb": "Ticker XTB",
                "ticker_yahoo": "Ticker Yahoo",
                "instrument": "Instrument",
                "open_time": "Otwarcie",
                "close_time": "Zamknięcie",
                "quantity": "Ilość",
                "open_price": "Cena otw.",
                "close_price": "Cena zamk.",
                "pnl": "PnL",
                "gross_pnl": "Gross PnL",
                "purchase_value": "Wartość zakupu",
                "sale_value": "Wartość sprzedaży",
            }
        )
        _write_dataframe(ws, cl, row, table_title="Zamknięte pozycje (XTB)")
    else:
        row = _write_section_title(ws, row, "Zamknięte pozycje (XTB)")
        ws.cell(row=row, column=1, value="Brak arkusza Closed Positions")

    _autosize_columns(ws)
    ws.freeze_panes = "A4"


def _summary_to_rows(summary: TradeAnalyticsSummary) -> list[tuple[str, object]]:
    pf = summary.profit_factor
    pf_val = f"{pf:.2f}" if pf < 100 else "∞"
    return [
        ("Round-tripy", summary.closed_trades),
        ("Win rate %", round(summary.win_rate_pct, 2)),
        ("Śr. czas trzymania (dni)", round(summary.avg_holding_days, 2)),
        ("Mediana trzymania (dni)", round(summary.median_holding_days, 2)),
        ("Śr. zysk", round(summary.avg_win, 2)),
        ("Śr. strata", round(summary.avg_loss, 2)),
        ("Profit factor", pf_val),
        ("Suma PnL zrealizowanego", round(summary.total_realized_pnl, 2)),
        ("Najlepszy trade", round(summary.best_trade_pnl, 2)),
        ("Najgorszy trade", round(summary.worst_trade_pnl, 2)),
    ]


def _build_analiza_sheet(
    ws,
    analyzed: pd.DataFrame,
    trade_summary: TradeAnalyticsSummary | None,
    cost_basis: pd.DataFrame | None,
    closed: pd.DataFrame | None,
    currency: str,
) -> None:
    ws.cell(row=1, column=1, value="Analiza").font = TITLE_FONT
    row = 3

    if trade_summary is not None and trade_summary.closed_trades > 0:
        row = _write_section_title(ws, row, "Trade analytics (FIFO)")
        row = _write_key_values(ws, row, _summary_to_rows(trade_summary))
    else:
        row = _write_section_title(ws, row, "Trade analytics (FIFO)")
        ws.cell(row=row, column=1, value="Brak round-tripów w Cash Operations")
        row += 2

    if closed is not None and not closed.empty:
        stats = closed_positions_summary(closed)
        row = _write_section_title(ws, row, "Zamknięte pozycje — statystyki XTB")
        row = _write_key_values(
            ws,
            row,
            [
                ("Liczba", stats["count"]),
                ("Łączny PnL", round(stats["total_pnl"], 2)),
                ("Win rate %", round(stats["win_rate_pct"], 2)),
                ("Zyskowne", stats["winners"]),
                ("Stratne", stats["losers"]),
            ],
        )

    if cost_basis is not None and not cost_basis.empty:
        current = get_current_cost_basis(cost_basis)
        if not current.empty:
            cb = current.rename(
                columns={
                    "ticker_xtb": "Ticker",
                    "ticker_yahoo": "Yahoo",
                    "quantity": "Ilość",
                    "avg_price": "Śr. cena",
                    "cost_basis": "Cost basis",
                    "last_trade_time": "Ostatnia transakcja",
                }
            )
            row = _write_dataframe(ws, cb, row, table_title="Cost basis (otwarte)")

    try:
        enriched = enrich_portfolio_allocation(analyzed)
        if not enriched.empty:
            sector = aggregate_breakdown(enriched, "sector").rename(
                columns={"sector": "Sektor", "market_value": f"Wartość ({currency})", "weight_pct": "Udział %"}
            )
            region = aggregate_breakdown(enriched, "region", sort_order=REGION_ORDER).rename(
                columns={"region": "Region", "market_value": f"Wartość ({currency})", "weight_pct": "Udział %"}
            )
            row = _write_dataframe(ws, sector, row, table_title="Alokacja sektorowa")
            _write_dataframe(ws, region, row, table_title="Alokacja geograficzna (USA/EU/PL)")
    except (ValueError, KeyError):
        row = _write_section_title(ws, row, "Alokacja")
        ws.cell(row=row, column=1, value="Nie udało się pobrać metadanych Yahoo")

    _autosize_columns(ws)
    ws.freeze_panes = "A4"


def generate_excel_bytes(
    report: XTBReport,
    analyzed: pd.DataFrame,
    summary: dict[str, float | str] | None = None,
) -> bytes:
    """
    Buduje workbook .xlsx w pamięci.

    Arkusze: Portfolio, Historia, Analiza.
    """
    if analyzed is None or analyzed.empty:
        raise ExcelExportError("Brak otwartych pozycji do eksportu.")

    summary = summary or portfolio_summary(analyzed)
    currency = str(summary["display_currency"])

    timeline: pd.DataFrame | None = None
    trades: pd.DataFrame | None = None
    round_trips: pd.DataFrame | None = None
    trade_summary: TradeAnalyticsSummary | None = None
    cost_basis: pd.DataFrame | None = None

    if report.cash_operations is not None:
        try:
            timeline = build_portfolio_timeline(report.cash_operations)
            trades = parse_cash_operations_trades(report.cash_operations)
            trade_summary, round_trips = compute_trade_analytics(
                trades, report.closed_positions
            )
            from core.cost_basis import build_cost_basis_history

            cost_basis = build_cost_basis_history(trades)
        except (ValueError, KeyError) as exc:
            raise ExcelExportError(f"Błąd analizy historii: {exc}") from exc

    closed = report.closed_positions

    wb = Workbook()
    ws_port = wb.active
    ws_port.title = "Portfolio"
    ws_hist = wb.create_sheet("Historia")
    ws_an = wb.create_sheet("Analiza")

    _build_portfolio_sheet(ws_port, report, analyzed, summary)
    _build_historia_sheet(ws_hist, report, timeline, trades, round_trips, closed)
    _build_analiza_sheet(ws_an, analyzed, trade_summary, cost_basis, closed, currency)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
